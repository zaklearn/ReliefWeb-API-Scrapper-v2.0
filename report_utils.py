"""
Utilitaires pour la génération de rapports avec crédits intégrés
"""
import pandas as pd
import io
from datetime import datetime
from credits import CREDITS_CONFIG, APP_HASH


def generate_excel_export(df: pd.DataFrame) -> bytes:
    """
    Convertit un DataFrame pandas en fichier Excel avec crédits intégrés.
    
    Args:
        df: DataFrame pandas à exporter
        
    Returns:
        Données binaires du fichier Excel
    """
    # Créer un buffer en mémoire
    output = io.BytesIO()
    
    # Écrire le DataFrame dans le buffer en format Excel
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Feuille principale avec les résultats
        df.to_excel(writer, index=False, sheet_name='Résultats')
        
        # Créer une feuille de crédits
        credits_data = {
            'Information': [
                'Application',
                'Version',
                'Généré le',
                'Développé par',
                'Site web',
                'Email',
                'Organisation',
                'Site organisation',
                'Licence',
                'Copyright',
                'App ID'
            ],
            'Valeur': [
                CREDITS_CONFIG['project_name'],
                CREDITS_CONFIG['version'],
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                CREDITS_CONFIG['author'],
                CREDITS_CONFIG['website'],
                CREDITS_CONFIG['contact'],
                CREDITS_CONFIG['organization'],
                CREDITS_CONFIG['org_website'],
                CREDITS_CONFIG['license'],
                f"© {CREDITS_CONFIG['year']}",
                APP_HASH
            ]
        }
        
        credits_df = pd.DataFrame(credits_data)
        credits_df.to_excel(writer, index=False, sheet_name='Crédits')
        
        # Ajuster la largeur des colonnes pour la feuille Résultats
        worksheet_results = writer.sheets['Résultats']
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(str(col))
            )
            # Limiter à 50 caractères max pour la largeur
            column_letter = chr(65 + idx) if idx < 26 else f"A{chr(65 + idx - 26)}"
            worksheet_results.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Ajuster la largeur des colonnes pour la feuille Crédits
        worksheet_credits = writer.sheets['Crédits']
        worksheet_credits.column_dimensions['A'].width = 25
        worksheet_credits.column_dimensions['B'].width = 50
        
        # Ajouter un style aux en-têtes (optionnel)
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Style pour les résultats
            for cell in worksheet_results[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Style pour les crédits
            for cell in worksheet_credits[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Mettre en gras la colonne Information dans Crédits
            for row in range(2, len(credits_data['Information']) + 2):
                worksheet_credits[f'A{row}'].font = Font(bold=True)
                
        except ImportError:
            pass  # Si openpyxl.styles n'est pas disponible, continuer sans style
    
    # Récupérer les données binaires
    output.seek(0)
    return output.getvalue()


def add_credits_footer_to_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Ajoute une ligne de crédits en bas du DataFrame (optionnel).
    
    Args:
        df: DataFrame original
        
    Returns:
        DataFrame avec ligne de crédits
    """
    # Créer une ligne vide
    empty_row = pd.DataFrame([[''] * len(df.columns)], columns=df.columns)
    
    # Créer une ligne de crédits
    credits_row = pd.DataFrame([[
        f"Généré par {CREDITS_CONFIG['project_name']} v{CREDITS_CONFIG['version']}",
        f"© {CREDITS_CONFIG['year']} {CREDITS_CONFIG['author']}",
        CREDITS_CONFIG['website'],
        '',
        '',
        ''
    ]], columns=df.columns)
    
    # Concaténer
    result = pd.concat([df, empty_row, credits_row], ignore_index=True)
    
    return result